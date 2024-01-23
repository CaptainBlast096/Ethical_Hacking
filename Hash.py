'''
Author: Jaleel Rogers
Class: CIS4204.01
Date: 01/25/24
'''
import openpyxl
import hashlib
import math

class Excel:
    def write_data(file_name, data):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Merging cells
        sheet.merge_cells('A1:C1')
        sheet.merge_cells('D1:G1')
        sheet.merge_cells('H1:O1')
        
        # Add Header
        sheet['A1'] = "Binary Value"
        sheet['D1'] = "MD5"
        sheet['H1'] = "SHA256"
        sheet['P1'] = "MD5 Decimal"
        sheet['Q1'] = "SHA256 Decimal"
        sheet['R1'] = "MD5 Difference"
        sheet['S1'] = "SHA256 Difference"
        sheet['T1'] = "MD5 Average Difference"
        sheet['U1'] = "SHA256 Average Difference"
        sheet['V1'] = "MD5 Standard Deviation"
        sheet['W1'] = "SHA256 Standard Deviation"

        # Arrays to hold hashes converter to decimal
        md5_decimals = []
        sha256_decimals = []
        
        # Add Data
        for row_index, value in enumerate(data, start=2):
            md5_hash = Hash.calculate_hash_md5(value)
            sha256_hash = Hash.calculate_sha256(value)
            md5_decimal = Hash.hash_to_decimal(md5_hash)
            sha256_decimal = Hash.hash_to_decimal(sha256_hash)

            sheet.cell(row=row_index, column=1, value=value)
            sheet.cell(row=row_index, column=4, value=md5_hash)
            sheet.cell(row=row_index, column=8, value=sha256_hash)
            sheet.cell(row=row_index, column=16, value=md5_decimal)
            sheet.cell(row=row_index, column=17, value=sha256_decimal)
            
            # Store MD5 Decimal values
            md5_decimals.append(md5_decimal) 
            md5_differences = [md5_decimals[i] - md5_decimals[i - 1] for i in range(1, len(md5_decimals))]
            for row_index, difference in enumerate(md5_differences, start=2):
                sheet.cell(row=row_index, column=18, value=difference)
                
            # Store SHA256 Decimal values   
            sha256_decimals.append(sha256_decimal) 
            sha256_differences = [sha256_decimals[i] - sha256_decimals[i - 1] for i in range(1, len(sha256_decimals))]
            for row_index, difference in enumerate(sha256_differences, start=2):
                sheet.cell(row=row_index, column=19, value=difference)
            
        # Calculate and add the average of MD5 Difference    
        md5_average_difference = Calculator.calculate_average(md5_differences)
        sheet['T2'] = md5_average_difference
        
        # Calculate and add the average of SHA256 Difference
        sha256_average_difference = Calculator.calculate_average(sha256_differences)
        sheet['U2'] = sha256_average_difference
            
        # Calculate and add the standard deviation of MD5 Difference    
        md5_standard_deviation = Calculator.calculate_standard_deviation(md5_differences)
        sheet['V2'] = md5_standard_deviation
        
        # Calculate and add the standard deviation of SHA256
        sha256_standard_deviation = Calculator.calculate_standard_deviation(sha256_differences)
        sheet['W2'] = sha256_standard_deviation
        
        # Save to file
        workbook.save(file_name)
        
# Class containing hashing algorithms and conversions
class Hash:
    def calculate_hash_md5(input_string):
        md5_hash = hashlib.md5(input_string.encode()).hexdigest()
        return md5_hash

    def calculate_sha256(input_string):
        sha256 = hashlib.sha256(input_string.encode()).hexdigest()
        return sha256

    def hash_to_decimal(hash_value):
        decimal_value = int(hash_value, 16)
        return decimal_value

# Class containing formulas
class Calculator:
    def calculate_average(data):
        average = sum(data) / len(data)
        return average
    
    def calculate_standard_deviation(data):
        mean = sum(data) / len(data)
        variance = sum((x - mean) ** 2 for x in data) / len(data)
        standard_deviation = math.sqrt(variance)
        return standard_deviation
    
def main():
    # Converts the integers to binary values and slices the last two values to only show the binary values
    binary_strings = [bin(x)[2:] for x in range(10000000, 10000300)]

    # File location
    excel_file = "D:\\binary_values.xlsx"

    excel = Excel()
    Excel.write_data(excel_file, binary_strings)

if __name__ == "__main__":
    main()
